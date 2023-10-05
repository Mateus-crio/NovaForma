import openpyxl
openpyxl.load_workbook("Dados.xlsx")
book = openpyxl.load_workbook("Dados.xlsx")

print(book.sheetnames) 
book.create_sheet("Custo de Produção")

dados_page = book["Custo de Produção"]

dados_page.append(["Valor", "Tempo"])
dados_page.append(["R$ 5.566.899", "12 Mês",])
dados_page.append(["R$ 139.543.156", "1 Mês",])
dados_page.append(["R$ 154.392.755", "2 Mês",])
dados_page.append(["R$ 504.367.921", "5 Mês",])

book.save("Dados.xlsx")