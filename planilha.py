import openpyxl
openpyxl.Workbook() 
book = openpyxl.Workbook()
print(book.sheetnames)
book.create_sheet("Fornecedores")

book["Fornecedores"]
Dados_page = book["Fornecedores"]

Dados_page.append(["Nome", "Telefone"])
Dados_page.append(["Mercedes São Paulo", "(11) 96546-7890"])
Dados_page.append(["Mercedes Santa Catarina", "(47) 78654-3210"])
Dados_page.append(["Mercedes Rio de Janeiro", "(21) 63654-3210"])


book.save("Dados.xlsx")